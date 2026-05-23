#!/usr/bin/env python3
"""
XLSM to Markdown Converter - Values & Formulas Version
Converts each sheet to TWO Markdown files:
- _values.md: Calculated/displayed values
- _formulas.md: Original Excel formulas
"""

import os
import sys
import zipfile
from pathlib import Path
from datetime import datetime

try:
    import openpyxl
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Error: openpyxl is required. Install with: pip install openpyxl")
    sys.exit(1)

def extract_vba_macros(xlsm_path, output_dir):
    """
    Extract VBA macro code from the xlsm file
    """
    macros_found = False
    macro_content = []
    
    try:
        with zipfile.ZipFile(xlsm_path, 'r') as zip_file:
            vba_files = [f for f in zip_file.namelist() 
                        if f.startswith('xl/vbaProject.bin') or f.endswith('.bin')]
            
            if vba_files:
                macro_content.append("# VBA Macros\n")
                macro_content.append("## Macro Information\n")
                macro_content.append(f"**File contains VBA macros** (detected {len(vba_files)} binary files)\n")
                macro_content.append("\n### Extracting Full Macro Code\n")
                macro_content.append("To extract the actual VBA code, use one of these methods:\n\n")
                macro_content.append("1. **Open in Excel**: Press `Alt+F11` to view/edit macros\n")
                macro_content.append("2. **Use oletools**: `python -m oletools.vba_extract original.xlsm > macros.vba`\n")
                macro_content.append("3. **Manual export**: In Excel VBA editor, right-click modules → Export File\n")
                macros_found = True
    except Exception as e:
        macro_content.append(f"Error reading macros: {str(e)}\n")
    
    if macros_found:
        macro_file = output_dir / "MACROS.md"
        with open(macro_file, 'w', encoding='utf-8') as f:
            f.write("\n".join(macro_content))
        return macro_file
    
    return None

def get_cell_display_value(cell):
    """
    Get the displayed value of a cell (what you see in Excel)
    """
    if cell.value is None:
        return ""
    
    # Handle different data types
    if isinstance(cell.value, (int, float)):
        if isinstance(cell.value, float) and cell.value.is_integer():
            return str(int(cell.value))
        return str(cell.value)
    elif isinstance(cell.value, datetime):
        return cell.value.strftime("%Y-%m-%d %H:%M:%S")
    else:
        return str(cell.value).replace('|', '\\|').replace('\n', ' ').strip()

def get_cell_formula(cell):
    """
    Get the formula from a cell if it exists, otherwise get the value
    """
    if cell.data_type == 'f' or (hasattr(cell, 'formula') and cell.formula):
        # Cell contains a formula
        formula = cell.formula if hasattr(cell, 'formula') else cell.value
        if formula:
            return f"`{formula}`"  # Format as code in Markdown
    elif cell.value is not None:
        # Static value
        return get_cell_display_value(cell)
    return ""

def sheet_to_markdown_values(worksheet, sheet_name):
    """
    Convert sheet to Markdown showing CALCULATED VALUES (what you see in Excel)
    """
    lines = []
    
    # Header
    lines.append(f"# Sheet: {sheet_name} - VALUES (Calculated/Displayed)\n")
    lines.append(f"*This shows the calculated values as they appear in Excel*\n")
    
    # Get used range
    max_row = worksheet.max_row
    max_col = worksheet.max_column
    
    if max_row == 1 and max_col == 1 and worksheet.cell(1, 1).value is None:
        lines.append("*This sheet appears to be empty*\n")
        return "\n".join(lines)
    
    # Limit for performance
    max_row = min(max_row, 5000)
    max_col = min(max_col, 50)
    
    # Find actual data range
    has_data = False
    for row in range(1, min(max_row + 1, 100)):
        for col in range(1, min(max_col + 1, 50)):
            if worksheet.cell(row, col).value is not None:
                has_data = True
                break
        if has_data:
            break
    
    if not has_data:
        lines.append("*This sheet appears to be empty*\n")
        return "\n".join(lines)
    
    # Create headers (first row)
    lines.append("## Data Table\n")
    
    headers = []
    for col in range(1, max_col + 1):
        header_value = worksheet.cell(1, col).value
        if header_value is None:
            header_value = f"Column_{get_column_letter(col)}"
        else:
            header_value = str(header_value).strip().replace('|', '\\|')[:100]
        headers.append(header_value)
    
    lines.append("| " + " | ".join(headers) + " |")
    lines.append("|" + "|".join([" --- " for _ in headers]) + "|")
    
    # Data rows (values)
    row_count = 0
    for row in range(2, max_row + 1):
        row_data = []
        row_has_data = False
        
        for col in range(1, max_col + 1):
            cell = worksheet.cell(row, col)
            value = get_cell_display_value(cell)
            
            if value:
                row_has_data = True
            row_data.append(value if value else "")
        
        if row_has_data:
            lines.append("| " + " | ".join(row_data) + " |")
            row_count += 1
    
    # Add metadata
    lines.append(f"\n---\n")
    lines.append(f"**Rows with data:** {row_count}")
    lines.append(f"**Columns:** {max_col}")
    lines.append(f"**File type:** Values (calculated results)")
    
    return "\n".join(lines)

def sheet_to_markdown_formulas(worksheet, sheet_name):
    """
    Convert sheet to Markdown showing FORMULAS (original Excel formulas)
    """
    lines = []
    
    # Header
    lines.append(f"# Sheet: {sheet_name} - FORMULAS (Original Excel Formulas)\n")
    lines.append(f"*This shows the actual formulas as entered in Excel*\n")
    lines.append(f"*Formulas are shown in `code blocks` for clarity*\n")
    
    # Get used range
    max_row = worksheet.max_row
    max_col = worksheet.max_column
    
    if max_row == 1 and max_col == 1 and worksheet.cell(1, 1).value is None:
        lines.append("*This sheet appears to be empty*\n")
        return "\n".join(lines)
    
    # Limit for performance
    max_row = min(max_row, 5000)
    max_col = min(max_col, 50)
    
    # Create headers (first row)
    lines.append("## Formula Table\n")
    
    headers = []
    for col in range(1, max_col + 1):
        header_value = worksheet.cell(1, col).value
        if header_value is None:
            header_value = f"Column_{get_column_letter(col)}"
        else:
            header_value = str(header_value).strip().replace('|', '\\|')[:100]
        headers.append(header_value)
    
    lines.append("| " + " | ".join(headers) + " |")
    lines.append("|" + "|".join([" --- " for _ in headers]) + "|")
    
    # Data rows (formulas)
    row_count = 0
    formula_count = 0
    
    for row in range(2, max_row + 1):
        row_data = []
        row_has_data = False
        
        for col in range(1, max_col + 1):
            cell = worksheet.cell(row, col)
            formula_value = get_cell_formula(cell)
            
            if formula_value:
                row_has_data = True
                if cell.data_type == 'f' or (hasattr(cell, 'formula') and cell.formula):
                    formula_count += 1
            row_data.append(formula_value if formula_value else "")
        
        if row_has_data:
            lines.append("| " + " | ".join(row_data) + " |")
            row_count += 1
    
    # Add metadata and legend
    lines.append(f"\n---\n")
    lines.append(f"## Summary\n")
    lines.append(f"**Rows with data:** {row_count}")
    lines.append(f"**Columns:** {max_col}")
    lines.append(f"**Cells containing formulas:** {formula_count}")
    
    lines.append(f"\n## Formula Legend\n")
    lines.append(f"- **`=FORMULA()`** : Excel formula (shown in code blocks)")
    lines.append(f"- **Plain text/number** : Static value (no formula)")
    lines.append(f"- **Empty cell** : No data")
    
    lines.append(f"\n## Tips\n")
    lines.append(f"- To copy a formula back to Excel, remove the backticks (`) and paste into a cell starting with `=`")
    lines.append(f"- Formulas are shown exactly as they appear in the Excel formula bar")
    
    return "\n".join(lines)

def convert_xlsm_to_markdown(xlsm_path, output_dir=None):
    """
    Main function - converts each sheet to VALUES and FORMULAS Markdown files
    """
    # Validate input
    xlsm_path = Path(xlsm_path)
    if not xlsm_path.exists():
        print(f"❌ Error: File '{xlsm_path}' not found")
        return False
    
    # Setup output directory
    if output_dir is None:
        output_dir = xlsm_path.stem + "_markdown_formulas"
    output_dir = Path(output_dir)
    output_dir.mkdir(exist_ok=True)
    
    print(f"📁 Output directory: {output_dir}")
    
    # Load workbook TWICE - once for values, once for formulas
    print(f"📖 Loading workbook: {xlsm_path.name}")
    
    try:
        # Load for values (data_only=True shows calculated values)
        wb_values = openpyxl.load_workbook(xlsm_path, data_only=True, keep_vba=True)
        print("   ✓ Loaded for VALUES extraction")
        
        # Load for formulas (data_only=False shows formulas)
        wb_formulas = openpyxl.load_workbook(xlsm_path, data_only=False, keep_vba=True)
        print("   ✓ Loaded for FORMULAS extraction")
    except Exception as e:
        print(f"❌ Error loading workbook: {str(e)}")
        return False
    
    # Extract macro info
    print("\n🔍 Extracting VBA macro information...")
    macro_file = extract_vba_macros(xlsm_path, output_dir)
    if macro_file:
        print(f"   ✓ Macro documentation: {macro_file.name}")
    
    # Create subdirectories
    values_dir = output_dir / "values"
    formulas_dir = output_dir / "formulas"
    values_dir.mkdir(exist_ok=True)
    formulas_dir.mkdir(exist_ok=True)
    
    # Process each sheet
    sheet_names = wb_values.sheetnames
    print(f"\n📝 Processing {len(sheet_names)} sheets...")
    
    values_files = []
    formulas_files = []
    
    for sheet_name in sheet_names:
        print(f"\n   📄 Sheet: {sheet_name}")
        
        # Create safe filename
        safe_name = "".join(c for c in sheet_name if c.isalnum() or c in (' ', '-', '_')).rstrip()
        safe_name = safe_name.replace(' ', '_')
        
        # Get worksheets from both workbooks
        ws_values = wb_values[sheet_name]
        ws_formulas = wb_formulas[sheet_name]
        
        # Generate VALUES markdown
        try:
            values_content = sheet_to_markdown_values(ws_values, sheet_name)
            values_path = values_dir / f"{safe_name}_values.md"
            with open(values_path, 'w', encoding='utf-8') as f:
                f.write(values_content)
            values_files.append(values_path)
            print(f"      ✓ VALUES: {values_path.name}")
        except Exception as e:
            print(f"      ❌ VALUES error: {str(e)}")
        
        # Generate FORMULAS markdown
        try:
            formulas_content = sheet_to_markdown_formulas(ws_formulas, sheet_name)
            formulas_path = formulas_dir / f"{safe_name}_formulas.md"
            with open(formulas_path, 'w', encoding='utf-8') as f:
                f.write(formulas_content)
            formulas_files.append(formulas_path)
            print(f"      ✓ FORMULAS: {formulas_path.name}")
        except Exception as e:
            print(f"      ❌ FORMULAS error: {str(e)}")
    
    # Create comprehensive index
    index_path = output_dir / "README.md"
    with open(index_path, 'w', encoding='utf-8') as f:
        f.write(f"# {xlsm_path.stem} - XLSM to Markdown Conversion\n\n")
        f.write(f"**Original file:** `{xlsm_path.name}`  \n")
        f.write(f"**Converted:** {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n\n")
        
        f.write("## Structure\n\n")
        f.write("This conversion creates **TWO versions** of each sheet:\n\n")
        f.write("- **VALUES** (`*_values.md`): Shows calculated/displayed values as seen in Excel\n")
        f.write("- **FORMULAS** (`*_formulas.md`): Shows original Excel formulas\n\n")
        
        f.write("## Sheets Converted\n\n")
        f.write("| Sheet Name | Values File | Formulas File |\n")
        f.write("|------------|-------------|---------------|\n")
        
        for sheet_name in sheet_names:
            safe_name = "".join(c for c in sheet_name if c.isalnum() or c in (' ', '-', '_')).rstrip()
            safe_name = safe_name.replace(' ', '_')
            f.write(f"| {sheet_name} | [VALUES](values/{safe_name}_values.md) | [FORMULAS](formulas/{safe_name}_formulas.md) |\n")
        
        if macro_file:
            f.write(f"\n## VBA Macros\n")
            f.write(f"- [Macro Documentation](MACROS.md)\n")
        
        f.write("\n## Directory Structure\n\n")
        f.write("```\n")
        f.write(f"{output_dir.name}/\n")
        f.write("├── README.md (this file)\n")
        f.write("├── MACROS.md (VBA documentation)\n")
        f.write("├── values/\n")
        f.write("│   ├── " + "\\n│   ├── ".join([f"{Path(f).name}" for f in values_files[:5]]) + "\n")
        if len(values_files) > 5:
            f.write(f"│   └── ... ({len(values_files)-5} more files)\n")
        else:
            f.write("│   └── ...\n")
        f.write("└── formulas/\n")
        f.write("    ├── " + "\\n    ├── ".join([f"{Path(f).name}" for f in formulas_files[:5]]) + "\n")
        if len(formulas_files) > 5:
            f.write(f"    └── ... ({len(formulas_files)-5} more files)\n")
        else:
            f.write("    └── ...\n")
        f.write("```\n")
        
        f.write("\n## Usage Tips\n\n")
        f.write("### Working with VALUES files\n")
        f.write("- These show the actual calculated results\n")
        f.write("- Numbers, dates, and text as they appear in Excel\n")
        f.write("- Great for documentation and reporting\n\n")
        
        f.write("### Working with FORMULAS files\n")
        f.write("- Formulas are shown in `code blocks` for visibility\n")
        f.write("- To reuse a formula in Excel: copy without backticks\n")
        f.write("- Static values (non-formulas) are shown as plain text\n\n")
        
        f.write("### Reconstructing in Excel\n")
        f.write("1. Create a new Excel file\n")
        f.write("2. Copy data from VALUES file for static data\n")
        f.write("3. Copy formulas from FORMULAS file (remove backticks)\n")
        f.write("4. Paste into cells starting with `=`\n")
    
    # Create summary
    summary_path = output_dir / "CONVERSION_SUMMARY.txt"
    with open(summary_path, 'w', encoding='utf-8') as f:
        f.write(f"XLSM Conversion Summary\n")
        f.write(f"=" * 50 + "\n\n")
        f.write(f"Original file: {xlsm_path.name}\n")
        f.write(f"File size: {xlsm_path.stat().st_size / 1024:.2f} KB\n")
        f.write(f"Number of sheets: {len(sheet_names)}\n")
        f.write(f"Output directory: {output_dir.absolute()}\n\n")
        f.write(f"Generated files:\n")
        f.write(f"  - Values files: {len(values_files)}\n")
        f.write(f"  - Formulas files: {len(formulas_files)}\n")
        if macro_file:
            f.write(f"  - Macro documentation: 1\n")
        f.write(f"\nConversion completed: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
    
    print("\n" + "="*50)
    print(f"✅ CONVERSION COMPLETE!")
    print(f"="*50)
    print(f"📊 Sheets processed: {len(sheet_names)}")
    print(f"📄 VALUES files: {len(values_files)}")
    print(f"🔣 FORMULAS files: {len(formulas_files)}")
    print(f"📁 Output directory: {output_dir.absolute()}")
    print(f"📖 Main index: {index_path}")
    print(f"📋 Summary: {summary_path}")
    
    return True

def main():
    """
    Command line interface
    """
    if len(sys.argv) < 2:
        print("XLSM to Markdown Converter - VALUES & FORMULAS version")
        print("=" * 55)
        print("\nUsage:")
        print("  python xlsm_to_markdown_vf.py <file.xlsm> [output_directory]")
        print("\nWhat it does:")
        print("  - Creates TWO Markdown files per sheet:")
        print("    1. *_values.md   - Calculated/displayed values")
        print("    2. *_formulas.md - Original Excel formulas")
        print("  - Extracts VBA macro documentation")
        print("  - Generates a complete index and summary")
        print("\nExamples:")
        print("  python xlsm_to_markdown_vf.py financial_model.xlsm")
        print("  python xlsm_to_markdown_vf.py data.xlsm ./output_folder")
        print("\nOptions:")
        print("  --help, -h      Show this help message")
        sys.exit(1)
    
    if sys.argv[1] in ['--help', '-h']:
        print(__doc__)
        sys.exit(0)
    
    xlsm_file = sys.argv[1]
    output_dir = sys.argv[2] if len(sys.argv) > 2 else None
    
    success = convert_xlsm_to_markdown(xlsm_file, output_dir)
    sys.exit(0 if success else 1)

if __name__ == "__main__":
    main()